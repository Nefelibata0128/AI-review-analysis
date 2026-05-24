"""
AI评论分析平台 v3.0 — HTTP 代理服务器
├─ 端口 8080，单文件静态服务 (report-dashboard.html)
├─ CORS 头自动注入
├─ Chrome UA 伪装 (绕过 Cloudflare Bot Detection)
├─ SSE 流式中继 (防 Cloudflare 504 超时)
├─ 路由分发:
│   ├─ /api/*        → Dify API (单工作流回退模式)
│   ├─ /api/orch/run    → 多 Agent 编排 (接收双文件上传)
│   └─ /api/orch/events → SSE 进度流
└─ .docx 文件解析 (python-docx)

用法: python proxy-server.py
"""
import http.server
import urllib.request
import urllib.error
import json
import ssl
import os
import sys
import cgi
import io
import threading
import queue
from pathlib import Path

PORT = 8091
DIFY_BASE = "https://api.dify.ai/v1"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"


def get_api_key():
    """从环境变量读取 Dify API Key"""
    return os.environ.get("DIFY_API_KEY", "")


def parse_docx(file_data: bytes) -> str:
    """解析 .docx 文件内容为纯文本"""
    from docx import Document
    doc = Document(io.BytesIO(file_data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def parse_xlsx(file_data: bytes) -> str:
    """解析 .xlsx/.xls 文件内容为 Markdown 表格文本"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        if len(wb.sheetnames) > 1:
            parts.append(f"## {sheet_name}")
        # 转为 Markdown 表格
        header = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("|" + "|".join(["---"] * len(header)) + "|")
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("| " + " | ".join(cells) + " |")
        parts.append("")
    wb.close()
    return "\n".join(parts)


def handle_file_upload(form: dict) -> tuple[str, str]:
    """解析 multipart 表单中的双文件。

    Returns:
        (review_text, bg_text): 评论文本和产品背景文本
    """
    review_text = ""
    bg_text = ""

    for key in ("review_file", "review_text"):
        if key in form:
            item = form[key]
            if hasattr(item, "file"):
                data = item.file.read()
                filename = (item.filename or "").lower()
                if filename.endswith(".docx"):
                    review_text = parse_docx(data)
                elif filename.endswith((".xlsx", ".xls")):
                    review_text = parse_xlsx(data)
                else:
                    review_text = data.decode("utf-8", errors="replace")
            elif isinstance(item, (str, bytes)):
                review_text = item.decode("utf-8", errors="replace") if isinstance(item, bytes) else item
            break

    for key in ("bg_file", "bg_text", "product_background"):
        if key in form:
            item = form[key]
            if hasattr(item, "file"):
                data = item.file.read()
                filename = (item.filename or "").lower()
                if filename.endswith(".docx"):
                    bg_text = parse_docx(data)
                elif filename.endswith((".xlsx", ".xls")):
                    bg_text = parse_xlsx(data)
                else:
                    bg_text = data.decode("utf-8", errors="replace")
            elif isinstance(item, (str, bytes)):
                bg_text = item.decode("utf-8", errors="replace") if isinstance(item, bytes) else item
            break

    return review_text, bg_text


class ProxyHandler(http.server.BaseHTTPRequestHandler):

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            html_path = os.path.join(os.path.dirname(__file__), "report-dashboard.html")
            if not os.path.exists(html_path):
                self._send(200, "<h1>AI评论分析平台 v3.0</h1><p>report-dashboard.html not found</p>".encode("utf-8"), "text/html; charset=utf-8")
                return
            with open(html_path, "r", encoding="utf-8") as f:
                html = f.read()
            self._send(200, html.encode("utf-8"), "text/html; charset=utf-8")

        elif self.path == "/api/orch/events":
            # SSE 进度流 — 从 orchestrator 的事件队列读取
            self._handle_sse_events()

        else:
            self.send_error(404)

    def do_POST(self):
        if self.path.startswith("/api/orch/run"):
            self._handle_orch_run()
        elif self.path.startswith("/api/"):
            self._handle_dify_proxy()
        else:
            self.send_error(404)

    def do_OPTIONS(self):
        self._send(200, b"", "text/plain")

    # ---- Dify 代理（保留 v2.0 逻辑） ----

    def _handle_dify_proxy(self):
        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length) if content_length > 0 else b""
        content_type = self.headers.get("Content-Type", "")

        dify_path = self.path.replace("/api", "/v1", 1)
        dify_url = DIFY_BASE.replace("/v1", "") + dify_path

        api_key = get_api_key()
        if not api_key:
            self._send(500, json.dumps({"error": "DIFY_API_KEY 未设置"}).encode(), "application/json")
            return

        try:
            req = urllib.request.Request(dify_url, data=body, method="POST")
            req.add_header("Authorization", f"Bearer {api_key}")
            req.add_header("Content-Type", content_type or "application/json")
            req.add_header("User-Agent", UA)
            req.add_header("Accept", "*/*")

            try:
                ctx = ssl.create_default_context()
            except Exception:
                ctx = ssl._create_unverified_context()
            resp = urllib.request.urlopen(req, timeout=300, context=ctx)
            resp_content_type = resp.headers.get("Content-Type", "")

            if "text/event-stream" in resp_content_type or "/workflows/run" in dify_path:
                self._stream_response(resp)
            else:
                self._send(resp.status, resp.read(), "application/json")

        except urllib.error.HTTPError as e:
            self._send(e.code, e.read(), "application/json")
        except Exception as e:
            self._send(500, json.dumps({"error": str(e)}).encode(), "application/json")

    # ---- 多 Agent 编排 ----

    def _handle_orch_run(self):
        """POST /api/orch/run — 接收双文件上传，启动多 Agent 分析"""
        content_type = self.headers.get("Content-Type", "")

        try:
            if "multipart/form-data" in content_type:
                form = cgi.parse(self.rfile, headers=self.headers, environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": content_type,
                })
                review_text, bg_text = handle_file_upload(form)
            else:
                content_length = int(self.headers.get("Content-Length", 0))
                body = self.rfile.read(content_length) if content_length > 0 else b"{}"
                data = json.loads(body.decode("utf-8", errors="replace"))
                review_text = data.get("review_text", "")
                bg_text = data.get("bg_text", "")

            mode = self._get_query_param("mode", "feedback")

            if not review_text.strip():
                self._send(400, json.dumps({"error": "review_text 为空"}).encode(), "application/json")
                return

            # 在后台线程中运行分析，结果通过 SSE 事件队列推送
            event_queue = queue.Queue()
            self.server.pending_events = event_queue

            thread = threading.Thread(
                target=self._run_analysis_thread,
                args=(review_text, bg_text, mode, event_queue),
                daemon=True,
            )
            thread.start()

            self._send(200, json.dumps({
                "status": "started",
                "message": "分析已启动，请连接 /api/orch/events 获取进度",
            }).encode(), "application/json")

        except Exception as e:
            self._send(500, json.dumps({"error": str(e)}).encode(), "application/json")

    def _run_analysis_thread(self, review_text: str, bg_text: str, mode: str, event_queue: queue.Queue):
        """后台线程：运行编排器，将事件推送到队列"""
        from orchestrator import Orchestrator

        orch = Orchestrator()
        # 替换 orchestrator 的 sse emitter，改为推送到共享队列
        orch.sse.event_queue = event_queue

        try:
            orch.run_analysis(review_text, bg_text, mode)
        except Exception as e:
            error_event = {
                "event": "error",
                "data": json.dumps({"agent": "Orchestrator", "message": str(e)}, ensure_ascii=False),
            }
            event_queue.put(error_event)
        finally:
            event_queue.put(None)  # 结束信号

    def _handle_sse_events(self):
        """GET /api/orch/events — SSE 流式推送编排进度"""
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Type", "text/event-stream")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Connection", "keep-alive")
        self.end_headers()

        event_queue = getattr(self.server, "pending_events", None)
        if event_queue is None:
            self.wfile.write(f"event: error\ndata: {json.dumps({'message': '无进行中的分析'})}\n\n".encode())
            self.wfile.flush()
            return

        heartbeat_count = 0
        try:
            while True:
                try:
                    item = event_queue.get(timeout=5)
                except queue.Empty:
                    heartbeat_count += 1
                    if heartbeat_count % 3 == 0:
                        self.wfile.write(": heartbeat\n\n".encode())
                        self.wfile.flush()
                    continue

                if item is None:
                    break

                if isinstance(item, str):
                    self.wfile.write(f"{item}\n\n".encode())
                elif isinstance(item, dict):
                    self.wfile.write(f"event: {item['event']}\ndata: {item['data']}\n\n".encode())
                self.wfile.flush()

        except (BrokenPipeError, ConnectionResetError):
            pass

    # ---- 通用方法 ----

    def _stream_response(self, upstream_resp):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Type", "text/event-stream")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Connection", "keep-alive")
        self.end_headers()

        try:
            while True:
                chunk = upstream_resp.read(4096)
                if not chunk:
                    break
                self.wfile.write(chunk)
                self.wfile.flush()
        except Exception as e:
            print(f"Stream error: {e}", file=sys.stderr)

    def _send(self, code, body, content_type):
        self.send_response(code)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _get_query_param(self, name: str, default: str = "") -> str:
        """从 URL query string 提取参数"""
        if "?" not in self.path:
            return default
        qs = self.path.split("?", 1)[1]
        for pair in qs.split("&"):
            if "=" in pair:
                k, v = pair.split("=", 1)
                if k == name:
                    return v
        return default

    def log_message(self, format, *args):
        print(f"[{self.log_date_time_string()}] {args[0]}")


if __name__ == "__main__":
    server = http.server.HTTPServer(("0.0.0.0", PORT), ProxyHandler)
    print(f"  AI评论分析平台 v3.0")
    print(f"  本地地址: http://localhost:{PORT}")
    print(f"  API 代理:    /api/* -> {DIFY_BASE}/*")
    print(f"  多Agent分析: /api/orch/run")
    print(f"  SSE 进度流:  /api/orch/events")
    print(f"  按 Ctrl+C 停止\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止")
